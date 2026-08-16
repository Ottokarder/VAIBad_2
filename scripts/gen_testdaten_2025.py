#!/usr/bin/env python3
"""Erzeugt aus daten/datei.xlsx die Datei vaibad_2_testdaten_2025.sql.

Liest die drei Stammdaten-Blaetter und schreibt INSERT-Befehle passend zum
Schema in vaibad_2_database_setup.sql.

WICHTIG zum Sponsoren-Modell (Stand: ohne Schema-Erweiterung):
In der Datenbank hat jeder Sponsor genau einen globalen betrag_pro_bahn und
ein globales limit. In der Excel sind diese Werte aber pro
Schwimmer-Sponsor-Paar unterschiedlich (derselbe Sponsor zahlt bei
Schwimmer A 0,50/Bahn und bei Schwimmer B 1,00/Bahn, evtl. mit Limit).

Da die Tabelle schwimmer_sponsor nur eine reine Zuordnungstabelle ist
(schwimmer_id, sponsoren_id), werden Sponsoren mit unterschiedlichen
Konditionen MEHRFACH angelegt: fuer jedes Schwimmer-Sponsor-Paar ein eigener
Datensatz in der Tabelle Sponsoren, mit demselben Namen, aber dem jeweiligen
betrag_pro_bahn und limit. Unterschieden werden die Datensaetze ueber die id.
Eine manuelle Zusammenfuehrung gleichnamiger Sponsoren ist spaeter moeglich.

Regeln (vom Nutzer vorgegeben):
- Schwimmer: Nr <= 50 -> nur vormittags geschwommen (Bahnen = vormittag, nachmittag=0).
  Nr > 50 -> nur nachmittags (Bahnen = nachmittag, vormittag=0).
  Eine Summe "a+a" -> erste Zahl vormittags, zweite nachmittags.
  Leere Bahnen -> 0/0.
- Geburtstag -> Geburtsjahr (Jahreszahl) in der DB.
- Mannschaft (Team-Nummer): kann mehrere Teams enthalten, getrennt durch Komma ODER Punkt
  (z.B. "503, 504" oder "503.505"). Schwimmer wird allen genannten Teams zugeordnet.
  Team 101/Platzhalter werden ignoriert. Schwimmer ohne Mannschaft = kein Team.
- Sponsoren: Name aus der Sponsoren-Stammdaten-Tabelle; betrag_pro_bahn und limit
  gelten PRO Schwimmer-Sponsor-Paar aus dem Schwimmer-Blatt.
  Der "Max. Betrag" aus der Sponsoren-Stammdaten-Tabelle ist bedeutungslos (wird ignoriert).
- Hauptsponsoren: keine anlegen.
- Teams: nur Teams mit echtem Namen anlegen (501-505). Platzhalter-Teams ab 506 leer -> weglassen.
- Team-Mitglieder: aus Schwimmer-Blatt (Mannschaft-Spalte) ableiten, NICHT aus Team-Blatt
  (das enthaellt 101-Platzhalter). Quelle der Wahrheit fuer die Zuordnung ist die
  Mannschaft-Spalte der Schwimmer.
"""
import openpyxl
import re

XLSX = "daten/datei.xlsx"
OUT = "vaibad_2_testdaten_2025.sql"


def sql_str(s):
    if s is None:
        return "NULL"
    s = str(s).strip()
    # Zeilenumbrueche und Steuerzeichen escapen, damit der String einzeilig
    # und SQL-syntaktisch sicher bleibt (z.B. TVV + Umbruch + Schwimmabteilung).
    s = s.replace("\\", "\\\\")
    s = s.replace("'", "''")
    s = s.replace("\n", "\\n")
    s = s.replace("\r", "\\r")
    s = s.replace("\t", "\\t")
    return f"'{s}'"


def to_int(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, int):
        return v
    m = re.search(r"\d+", str(v))
    return int(m.group()) if m else None


def parse_bahnen(bahnen, nr):
    """Gibt (vormittag, nachmittag) zurueck."""
    if bahnen is None:
        return (0, 0)
    if isinstance(bahnen, (int, float)):
        b = int(bahnen)
        if nr is not None and nr <= 50:
            return (b, 0)
        return (0, b)
    # String: moegliche Summe "a+a"
    s = str(bahnen).replace(" ", "")
    if "+" in s:
        parts = s.split("+")
        try:
            v = int(parts[0]) if parts[0] else 0
            n = int(parts[1]) if len(parts) > 1 and parts[1] else 0
            return (v, n)
        except ValueError:
            pass
    # einfache Zahl als String
    m = re.search(r"\d+", s)
    if m:
        b = int(m.group())
        if nr is not None and nr <= 50:
            return (b, 0)
        return (0, b)
    return (0, 0)


def parse_teams(mannschaft):
    """Gibt Liste von Team-Nummern zurueck."""
    if mannschaft is None:
        return []
    s = str(mannschaft)
    # trenner: komma oder punkt
    parts = re.split(r"[,;.]", s)
    teams = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.search(r"\d+", p)
        if m:
            t = int(m.group())
            if t != 101:  # 101 = Platzhalter "kein Eintrag"
                teams.append(t)
    return teams


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ---- Sponsor-Namen einlesen (Nr -> Name) ----
    # Nur der Name interessiert; betrag_pro_bahn und limit werden pro Paar
    # aus dem Schwimmer-Blatt bezogen.
    ws_sp = wb["Stammdaten - Sponsoren"]
    sponsor_name = {}  # nr -> name (firma + vertreter)
    for r in range(4, ws_sp.max_row + 1):
        nr = ws_sp.cell(r, 1).value
        name = ws_sp.cell(r, 2).value
        vertreter = ws_sp.cell(r, 3).value
        if nr is None:
            continue
        nr = to_int(nr)
        if nr is None:
            continue
        if name is None and vertreter is None:
            continue  # leerer Eintrag
        # Name zusammenbauen: "Firma - Vertreter" oder nur Firma oder nur Vertreter
        parts = []
        if name:
            parts.append(str(name).strip())
        if vertreter:
            parts.append(str(vertreter).strip())
        full_name = " - ".join(parts) if len(parts) == 2 else parts[0]
        sponsor_name[nr] = full_name

    # ---- Teams einlesen (nur echte Teams mit Namen) ----
    ws_tm = wb["Stammdaten Teams"]
    teams = {}  # nr -> dict(name, betrag_pro_bahn)
    for r in range(4, ws_tm.max_row + 1):
        nr = ws_tm.cell(r, 1).value
        name = ws_tm.cell(r, 2).value
        betrag = ws_tm.cell(r, 6).value  # Spalte F = Spende pro Bahn
        if nr is None:
            continue
        nr = to_int(nr)
        if nr is None:
            continue
        if name is None:
            continue  # Platzhalter-Team ohne Namen -> ueberspringen
        # betrag_pro_bahn: Default 0.50 falls nicht gesetzt
        if betrag is None:
            betrag = 0.50
        teams[nr] = {"name": str(name).strip(), "betrag": float(betrag)}

    # ---- Schwimmer einlesen ----
    ws_sw = wb["Stammdaten - Schwimmer"]
    schwimmer = []  # list of dict
    for r in range(4, ws_sw.max_row + 1):
        nr = ws_sw.cell(r, 1).value
        if nr is None:
            continue
        nr = to_int(nr)
        if nr is None:
            continue
        name = ws_sw.cell(r, 2).value  # Nachname
        vorname = ws_sw.cell(r, 3).value
        if name is None and vorname is None:
            continue  # leerer Schwimmer
        gb = ws_sw.cell(r, 4).value
        if hasattr(gb, "year"):
            jahr = gb.year
        else:
            j = to_int(gb)
            jahr = j if j else 0
        mannschaft = ws_sw.cell(r, 5).value
        bahnen = ws_sw.cell(r, 6).value
        vm, nm = parse_bahnen(bahnen, nr)
        team_list = parse_teams(mannschaft)
        # Sponsor-Bloecke: G,H,I,J | K,L,M,N | O,P,Q,R | S,T,U,V
        sponsoren_list = []
        for start in (7, 11, 15, 19):
            snr = ws_sw.cell(r, start).value
            spb = ws_sw.cell(r, start + 1).value  # spendehoehe pro bahn
            mx = ws_sw.cell(r, start + 2).value   # max betrag (limit) pro paar
            if snr is None:
                continue
            snr_i = to_int(snr)
            if snr_i is None:
                continue
            if spb is None:
                continue  # ohne betrag pro bahn kein gueltiger eintrag
            try:
                spb_f = float(spb)
            except (ValueError, TypeError):
                continue
            mx_i = to_int(mx) if mx is not None else None
            sponsoren_list.append({
                "sponsoren_nr": snr_i,
                "betrag_pro_bahn": spb_f,
                "limit": mx_i,
            })
        schwimmer.append({
            "nr": nr,
            "nachname": str(name).strip() if name else "",
            "vorname": str(vorname).strip() if vorname else "",
            "geburtsjahr": jahr,
            "vm": vm,
            "nm": nm,
            "teams": team_list,
            "sponsoren": sponsoren_list,
        })

    # ---- SQL schreiben ----
    lines = []
    lines.append("-- Testdaten fuer vaibad_2 (Sponsorenschwimmen 2025)")
    lines.append("-- Automatisch erzeugt aus daten/datei.xlsx")
    lines.append("--")
    lines.append("-- Regeln:")
    lines.append("--   * Schwimmer Nr<=50: nur vormittags geschwommen.")
    lines.append("--   * Schwimmer Nr>50:  nur nachmittags geschwommen.")
    lines.append("--   * Summe 'a+a' in Bahnen-Spalte: a vormittags + a nachmittags.")
    lines.append("--   * Sponsoren werden PRO Schwimmer-Sponsor-Paar angelegt:")
    lines.append("--     gleicher Name, aber eigener betrag_pro_bahn und limit pro Paar.")
    lines.append("--     Unterschieden wird ueber die id. Eine manuelle Zusammenfuehrung")
    lines.append("--     gleichnamiger Sponsoren ist spaeter moeglich.")
    lines.append("--   * Keine Hauptsponsoren.")
    lines.append("--   * Geburtsjahr aus dem Geburtstag-Datum abgeleitet.")
    lines.append("--")
    lines.append("-- Ausfuehren mit:")
    lines.append("--   mysql -u root vaibad_2 < vaibad_2_testdaten_2025.sql")
    lines.append("-- Setzt eine bereits angelegte Datenbank vaibad_2 voraus")
    lines.append("-- (siehe vaibad_2_database_setup.sql).")
    lines.append("")
    lines.append("USE vaibad_2;")
    lines.append("")
    lines.append("-- Bestehende Testdaten loeschen (idempotent, in abhaengiger Reihenfolge).")
    lines.append("DELETE FROM spenden_hauptsponsoren;")
    lines.append("DELETE FROM spenden_teams;")
    lines.append("DELETE FROM spenden_sponsoren;")
    lines.append("DELETE FROM schwimmer_team;")
    lines.append("DELETE FROM schwimmer_sponsor;")
    lines.append("DELETE FROM Schwimmer;")
    lines.append("DELETE FROM Teams;")
    lines.append("DELETE FROM Sponsoren;")
    lines.append("DELETE FROM Hauptsponsoren;")
    lines.append("")
    lines.append("-- AUTO_INCREMENT zuruecksetzen, damit IDs sauber bei 1 beginnen.")
    lines.append("ALTER TABLE Schwimmer AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE Sponsoren AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE Teams AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE schwimmer_sponsor AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE schwimmer_team AUTO_INCREMENT = 1;")
    lines.append("")

    # --- Schwimmer einfuegen (zuerst, damit die Sponsor-Paare referenzieren koennen) ---
    lines.append(f"-- Schwimmer ({len(schwimmer)} Eintraege)")
    lines.append("INSERT INTO Schwimmer (startnummer, vorname, nachname, geburtsjahr, schwimmleistung_vormittag, schwimmleistung_nachmittag) VALUES")
    sw_rows = []
    for s in schwimmer:
        sw_rows.append(
            f"  ({s['nr']}, {sql_str(s['vorname'])}, {sql_str(s['nachname'])}, "
            f"{s['geburtsjahr']}, {s['vm']}, {s['nm']})"
        )
    lines.append(",\n".join(sw_rows) + ";")
    lines.append("")
    # Map startnummer -> DB-id (Schwimmer.id = fortlaufend ab 1 in Einfuege-Reihefolge)
    schwimmer_id_map = {}
    for idx, s in enumerate(schwimmer, start=1):
        schwimmer_id_map[s["nr"]] = idx

    # --- Sponsoren einfuegen: PRO Schwimmer-Sponsor-Paar ein eigener Datensatz ---
    # Name stammt aus der Sponsoren-Stammdaten-Tabelle; betrag_pro_bahn und limit
    # aus dem jeweiligen Schwimmer-Blatt-Eintrag. Sponsoren mit unterschiedlichen
    # Konditionen werden mehrfach (gleicher Name, unterschiedliche id) angelegt.
    sponsor_rows = []  # list of (name, betrag_pro_bahn, limit)
    verknuepfungen = []  # list of (schwimmer_id, sponsoren_id)
    missing_names = set()
    for s in schwimmer:
        sid = schwimmer_id_map[s["nr"]]
        # Doppelte Sponsor-Zuordnungen pro Schwimmer bereinigen (z.B. Nr.18 hat 1015 doppelt)
        seen = set()
        for sp in s["sponsoren"]:
            key = sp["sponsoren_nr"]
            if key in seen:
                continue
            seen.add(key)
            name = sponsor_name.get(sp["sponsoren_nr"])
            if name is None:
                missing_names.add(sp["sponsoren_nr"])
                continue  # Sponsor ohne Stammdaten-Eintrag -> skippen
            sponsor_rows.append((name, sp["betrag_pro_bahn"], sp["limit"]))
            sponsoren_id = len(sponsor_rows)  # AUTO_INCREMENT ab 1
            verknuepfungen.append((sid, sponsoren_id))

    lines.append(f"-- Sponsoren ({len(sponsor_rows)} Eintraege)")
    lines.append("-- Pro Schwimmer-Sponsor-Paar ein eigener Datensatz (gleicher Name moeglich,")
    lines.append("-- unterschieden ueber die id). betrag_pro_bahn/limit aus dem Schwimmer-Blatt.")
    lines.append("INSERT INTO Sponsoren (name, betrag_pro_bahn, `limit`) VALUES")
    sp_rows = []
    for (name, spb, mx) in sponsor_rows:
        lim = "NULL" if mx is None else str(mx)
        sp_rows.append(f"  ({sql_str(name)}, {spb:.2f}, {lim})")
    lines.append(",\n".join(sp_rows) + ";")
    lines.append("")

    # --- schwimmer_sponsor einfuegen: reine Zuordnungstabelle (nur IDs) ---
    if verknuepfungen:
        lines.append(f"-- Schwimmer-Sponsor Verknuepfungen ({len(verknuepfungen)} Eintraege)")
        lines.append("-- Reine Zuordnungstabelle; Betrag/Limit liegen am Sponsor.")
        lines.append("INSERT INTO schwimmer_sponsor (schwimmer_id, sponsoren_id) VALUES")
        vs_rows = [f"  ({sid}, {spon_id})" for (sid, spon_id) in verknuepfungen]
        lines.append(",\n".join(vs_rows) + ";")
        lines.append("")

    # --- Teams einfuegen ---
    lines.append(f"-- Teams ({len(teams)} Eintraege)")
    lines.append("INSERT INTO Teams (name, betrag_pro_bahn, `limit`) VALUES")
    tm_rows = []
    for nr in sorted(teams):
        t = teams[nr]
        # limit fuer Teams: nicht aus Excel ableitbar -> NULL
        tm_rows.append(f"  ({sql_str(t['name'])}, {t['betrag']:.2f}, NULL)")
    lines.append(",\n".join(tm_rows) + ";")
    lines.append("")
    team_id_map = {}
    for idx, nr in enumerate(sorted(teams), start=1):
        team_id_map[nr] = idx

    # --- schwimmer_team einfuegen ---
    tpairs = []
    seen_t = set()
    for s in schwimmer:
        sid = schwimmer_id_map[s["nr"]]
        for tnr in s["teams"]:
            tid = team_id_map.get(tnr)
            if tid is None:
                continue
            key = (sid, tid)
            if key in seen_t:
                continue
            seen_t.add(key)
            tpairs.append((sid, tid))
    if tpairs:
        lines.append(f"-- Schwimmer-Team Verknuepfungen ({len(tpairs)} Eintraege)")
        lines.append("INSERT INTO schwimmer_team (schwimmer_id, team_id) VALUES")
        vt_rows = [f"  ({sid}, {tid})" for (sid, tid) in tpairs]
        lines.append(",\n".join(vt_rows) + ";")
        lines.append("")

    lines.append("-- Ergebnistabellen (spenden_*) werden NICHT hier befuellt.")
    lines.append("-- Sie werden ueber die Spendenberechnung in der Webapp gefuellt")
    lines.append("-- (siehe webapp/spendenberechnung.php).")
    lines.append("")
    lines.append("SELECT 'Testdaten 2025 erfolgreich eingefuegt.' AS Erfolg;")
    lines.append("")
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # Statistik auf stderr
    import sys
    print(f"Sponsoren (Paar-Datensaetze): {len(sponsor_rows)}", file=sys.stderr)
    print(f"Teams: {len(teams)}", file=sys.stderr)
    print(f"Schwimmer: {len(schwimmer)}", file=sys.stderr)
    print(f"Schwimmer-Sponsor-Paare: {len(verknuepfungen)}", file=sys.stderr)
    print(f"Schwimmer-Team-Paare: {len(tpairs)}", file=sys.stderr)
    if missing_names:
        print(f"WARN: Sponsoren-Nrn ohne Eintrag in Sponsoren-Tabelle: {sorted(missing_names)}", file=sys.stderr)
    # gleichnamige Sponsoren (Hinweis fuer manuelle Zusammenfuehrung)
    from collections import Counter
    dup = {n: c for n, c in Counter(n for (n, _, _) in sponsor_rows).items() if c > 1}
    if dup:
        print(f"Hinweis: {len(dup)} Sponsor-Name(n) mehrfach angelegt (Zusammenfuehrung moeglich).", file=sys.stderr)
        for n, c in sorted(dup.items(), key=lambda kv: -kv[1])[:10]:
            print(f"   {c}x  {n}", file=sys.stderr)
    all_tm_used = set(t for s in schwimmer for t in s["teams"])
    missing_t = all_tm_used - set(team_id_map)
    if missing_t:
        print(f"WARN: Team-Nrn ohne Eintrag in Teams-Tabelle: {sorted(missing_t)}", file=sys.stderr)


if __name__ == "__main__":
    main()
